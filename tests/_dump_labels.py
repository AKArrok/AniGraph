"""Dump the human_correct column from human_labeling.xlsx to verify what
is actually saved on disk."""
from pathlib import Path
from openpyxl import load_workbook

wb = load_workbook(Path(__file__).parent / "human_labeling.xlsx")
ws = wb.active
header = [c.value for c in ws[1]]
idx = header.index("human_correct (fill 0/1)")
for row in ws.iter_rows(min_row=2, values_only=True):
    print(row[0], "->", repr(row[idx]))
