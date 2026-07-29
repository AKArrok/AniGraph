"""Rebuild tests/ablation_results.xlsx as two hard-set-only sheets.

Sheet 1 `Ablation`
  Rows: Overall Acc, per-type acc (6 types), Avg Latency, Avg Tokens.
  Cols: H0 / H1 / H2 / H3 / H4 / B5.

Sheet 2 `Cases`
  One row per case; per-experiment PASS/FAIL + truncated answer.

Everything is computed from tests/hard_eval_results.json and
tests/full_pipeline_results.json. Easy-set sheets are dropped.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
HARD_PATH = ROOT / "tests" / "hard_eval_results.json"
B5_PATH = ROOT / "tests" / "full_pipeline_results.json"
OUT = ROOT / "tests" / "ablation_results.xlsx"

HEADER = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PARTIAL_FILL = PatternFill("solid", fgColor="FFEB9C")
BORDER = Border(*[Side(style="thin", color="B4B4B4")] * 4)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

QUERY_TYPES = [
    "metadata_cross",
    "longtail_fact",
    "similar_recommendation",
    "bangumi_tags",
    "bangumi_score_precise",
    "kb_boundary",
]

EXPERIMENTS = [
    ("H0", "no RAG"),
    ("H1", "Dense k=20"),
    ("H2", "Dense k=3"),
    ("H3", "Dense k=10"),
    ("H5", "Dense k=40"),
    ("H6", "Dense k=60"),
    ("H4", "k=20 no-cite"),
    ("A1", "no fast path"),
    ("A2", "no eval conflict"),
    ("A3", "no query rewrite"),
    ("B5", "Full pipeline"),
]


def style_header(row):
    for cell in row:
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def aggregate(rows: list[dict]) -> dict:
    if not rows:
        return {}
    n = len(rows)
    correct = sum(int(r.get("correct", False)) for r in rows)
    by_type: dict[str, dict] = {}
    lat_sum = 0.0
    tok_sum = 0
    tok_n = 0
    for r in rows:
        b = by_type.setdefault(r["type"], {"n": 0, "correct": 0})
        b["n"] += 1
        b["correct"] += int(r.get("correct", False))
        lat_sum += r.get("latency", 0) or 0
        t = r.get("tokens")
        if t:
            tok_sum += t
            tok_n += 1
    for b in by_type.values():
        b["acc"] = b["correct"] / b["n"]
    return {
        "n": n,
        "acc": correct / n,
        "by_type": by_type,
        "avg_latency": lat_sum / n,
        "avg_tokens": (tok_sum / tok_n) if tok_n else None,
    }


def load_experiments() -> dict[str, dict]:
    hard = json.loads(HARD_PATH.read_text(encoding="utf-8"))
    b5 = json.loads(B5_PATH.read_text(encoding="utf-8"))

    exps: dict[str, dict] = {}
    for tag in ("H0", "H1", "H2", "H3", "H4", "H5", "H6"):
        rows = hard.get(f"details_{tag.lower()}", [])
        exps[tag] = {"agg": aggregate(rows), "rows": rows}

    # Component ablation (A1/A2/A3)
    for tag in ("A1", "A2", "A3"):
        comp_path = ROOT / "tests" / f"component_ablation_{tag}_results.json"
        if comp_path.exists():
            comp = json.loads(comp_path.read_text(encoding="utf-8"))
            rows = comp.get("details", [])
            exps[tag] = {"agg": aggregate(rows), "rows": rows}
        else:
            exps[tag] = {"agg": {}, "rows": []}

    b5_rows = b5.get("details", [])
    exps["B5"] = {"agg": aggregate(b5_rows), "rows": b5_rows}
    return exps


def color_acc(cell, v):
    if not isinstance(v, (int, float)):
        return
    if v >= 0.6:
        cell.fill = PASS_FILL
    elif v >= 0.3:
        cell.fill = PARTIAL_FILL
    else:
        cell.fill = FAIL_FILL


def build(exps: dict[str, dict]) -> None:
    wb = Workbook()
    del wb[wb.sheetnames[0]]

    # ── Sheet 1: Ablation ────────────────────────────────────────────
    ws = wb.create_sheet("Ablation")
    tags = [t for t, _ in EXPERIMENTS]
    ws.append(["Metric"] + tags)
    style_header(ws[1])
    ws.append(["Config"] + [d for _, d in EXPERIMENTS])
    for c in ws[2]:
        c.font = Font(italic=True, color="6C6C6C")
        c.alignment = CENTER
        c.border = BORDER

    def add(label, values, pct=True, fmt=None):
        ws.append([label] + list(values))
        row = ws.max_row
        for i, v in enumerate(values, 2):
            cell = ws.cell(row=row, column=i)
            if isinstance(v, float):
                if fmt:
                    cell.number_format = fmt
                elif pct:
                    cell.number_format = "0.0%"
                    color_acc(cell, v)
                else:
                    cell.number_format = "0.00"
        for cell in ws[row]:
            cell.border = BORDER
            cell.alignment = CENTER

    add("Overall Accuracy", [exps[t]["agg"].get("acc", "-") for t in tags])
    for qt in QUERY_TYPES:
        vals = []
        for t in tags:
            b = exps[t]["agg"].get("by_type", {}).get(qt)
            vals.append(b["acc"] if b else "-")
        add(qt, vals)
    add("Avg Latency (s)", [exps[t]["agg"].get("avg_latency", "-") for t in tags], pct=False)
    add(
        "Avg Tokens",
        [exps[t]["agg"].get("avg_tokens") or "-" for t in tags],
        pct=False,
        fmt="0",
    )

    ws.freeze_panes = "B3"
    col_widths(ws, [28] + [14] * len(tags))

    # ── Sheet 2: Cases ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Cases")
    headers = ["ID", "Type", "Query"] + [f"{t}" for t in tags]
    ws2.append(headers)
    style_header(ws2[1])

    # Collect IDs in stable order (first non-empty rows)
    ids: list[str] = []
    id_query: dict[str, tuple[str, str]] = {}
    for t in tags:
        for r in exps[t]["rows"]:
            if r["id"] not in id_query:
                id_query[r["id"]] = (r["type"], r["query"])
                ids.append(r["id"])

    row_index: dict[str, dict[str, dict]] = {
        t: {r["id"]: r for r in exps[t]["rows"]} for t in tags
    }

    for cid in ids:
        qtype, q = id_query[cid]
        row = [cid, qtype, q]
        for t in tags:
            r = row_index[t].get(cid)
            row.append("PASS" if r and r.get("correct") else ("FAIL" if r else "-"))
        ws2.append(row)
        rn = ws2.max_row
        for cell in ws2[rn]:
            cell.border = BORDER
            cell.alignment = LEFT if cell.column == 3 else CENTER
        for i, t in enumerate(tags):
            cell = ws2.cell(row=rn, column=4 + i)
            if cell.value == "PASS":
                cell.fill = PASS_FILL
            elif cell.value == "FAIL":
                cell.fill = FAIL_FILL

    ws2.freeze_panes = "D2"
    col_widths(ws2, [10, 22, 55] + [10] * len(tags))

    wb.save(OUT)


def main() -> None:
    exps = load_experiments()
    build(exps)
    print(f"Saved: {OUT}")
    for tag, desc in EXPERIMENTS:
        agg = exps[tag]["agg"]
        if not agg:
            print(f"  {tag:3} {desc:16} — no data")
            continue
        tok = agg.get("avg_tokens")
        tok_str = f"{tok:.0f}" if tok else "-"
        print(
            f"  {tag:3} {desc:16}  acc={agg['acc']:.1%}  "
            f"lat={agg['avg_latency']:.1f}s  tok={tok_str}"
        )


if __name__ == "__main__":
    main()
