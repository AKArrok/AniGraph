import json, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "tests", "ablation_results.xlsx")
H = json.load(open(os.path.join(ROOT, "tests", "hard_eval_results.json"), encoding="utf-8"))
B = json.load(open(os.path.join(ROOT, "tests", "full_pipeline_results.json"), encoding="utf-8"))

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PARTIAL_FILL = PatternFill("solid", fgColor="FFEB9C")
BORDER = Border(*[Side(style="thin", color="B4B4B4")]*4)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = load_workbook(XLSX)
for s in list(wb.sheetnames):
    if s.startswith("B5"):
        del wb[s]
ws = wb.create_sheet("B5 vs H1", 0)

qtypes = ["metadata_cross", "longtail_fact", "similar_recommendation",
          "bangumi_tags", "bangumi_score_precise", "kb_boundary"]
h1 = H["summary"]["h1"]
b5 = B["summary"]

ws.append(["Query Type", "H1 dense k=20", "B5 full pipeline", "Delta"])
for c in ws[1]:
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER

def color(cell, v):
    if v >= 0.6: cell.fill = PASS_FILL
    elif v >= 0.3: cell.fill = PARTIAL_FILL
    else: cell.fill = FAIL_FILL

for qt in qtypes:
    ha = h1["by_type"].get(qt, {}).get("acc", 0)
    ba = b5["by_type"].get(qt, {}).get("acc", 0)
    ws.append([qt, ha, ba, ba - ha])
    r = ws.max_row
    for col in (2, 3):
        cell = ws.cell(row=r, column=col)
        cell.number_format = "0.0%"; color(cell, cell.value)
    dcell = ws.cell(row=r, column=4)
    dcell.number_format = "+0.0%;-0.0%"
    if dcell.value < 0: dcell.fill = FAIL_FILL
    elif dcell.value > 0: dcell.fill = PASS_FILL
    for cell in ws[r]:
        cell.border = BORDER; cell.alignment = CENTER

ws.append(["OVERALL", h1["strict_accuracy"], b5["strict_accuracy"],
           b5["strict_accuracy"] - h1["strict_accuracy"]])
r = ws.max_row
for cell in ws[r]:
    cell.font = Font(bold=True); cell.border = BORDER; cell.alignment = CENTER
ws.cell(row=r, column=2).number_format = "0.0%"
ws.cell(row=r, column=3).number_format = "0.0%"
ws.cell(row=r, column=4).number_format = "+0.0%;-0.0%"

for i, w in enumerate([26, 16, 18, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# case-level B5 detail sheet
ws2 = wb.create_sheet("B5 Cases")
ws2.append(["ID", "Type", "H1", "B5", "Web?", "Latency(s)", "B5 Answer (trunc)"])
for c in ws2[1]:
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
h1_by_id = {r["id"]: r for r in H["details_h1"]}
for r in B["details"]:
    h1c = "PASS" if h1_by_id.get(r["id"], {}).get("correct") else "FAIL"
    b5c = "PASS" if r["correct"] else "FAIL"
    ws2.append([r["id"], r["type"], h1c, b5c, "Y" if r["web_used"] else "",
                round(r["latency"], 1), (r["answer"] or "")[:160]])
    rr = ws2.max_row
    for col, val in ((3, h1c), (4, b5c)):
        cell = ws2.cell(row=rr, column=col)
        cell.fill = PASS_FILL if val == "PASS" else FAIL_FILL
    for cell in ws2[rr]:
        cell.border = BORDER
        cell.alignment = LEFT if cell.column == 7 else CENTER
for i, w in enumerate([10, 22, 8, 8, 6, 11, 70], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save(XLSX)
print("saved", XLSX)
print("sheets:", wb.sheetnames)
print("\nH1 overall {:.1%}  vs  B5 overall {:.1%}".format(
    h1["strict_accuracy"], b5["strict_accuracy"]))
