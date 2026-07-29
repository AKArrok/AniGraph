"""Build ablation_results.xlsx from ablation_results.json."""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "tests", "ablation_results.json")
OUT = os.path.join(ROOT, "tests", "ablation_results.xlsx")

with open(SRC, encoding="utf-8") as f:
    data = json.load(f)

wb = Workbook()

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PARTIAL_FILL = PatternFill("solid", fgColor="FFEB9C")
BORDER = Border(*[Side(style="thin", color="B4B4B4")]*4)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(row):
    for cell in row:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 1: Summary comparison ──
ws = wb.active
ws.title = "Summary"
headers = ["Experiment", "N", "Accuracy", "Faithfulness", "Avg Latency (s)", "P95 Latency (s)", "Avg Tokens", "Notes"]
ws.append(headers)
style_header(ws[1])

EXPERIMENT_META = {
    "B0_direct_llm":   ("B0: Direct LLM",           "No RAG, LLM knowledge only"),
    "B1_dense_rag":    ("B1: Dense RAG",             "Pinecone dense retrieval + single LLM"),
    "B2_dense_topk3":  ("B2: Dense top_k=3",         "Dense RAG, tight retrieval"),
    "B3_dense_topk10": ("B3: Dense top_k=10",        "Dense RAG, mid retrieval"),
    "B4_no_cite_prompt":("B4: no-cite prompt",       "Dense top_k=20, prompt without 'cite sources'"),
    "B5_full":         ("B5: AniGraph Full",         "+ Evaluator + Replan + HyDE (pending)"),
    "AB_sparse":       ("Ablation: no Sparse",       "Full minus Whoosh"),
    "AB_reranker":     ("Ablation: no Reranker",     "Full minus CrossEncoder"),
    "AB_hyde":         ("Ablation: no HyDE",         "Full minus HyDE query rewrite"),
    "AB_multiexpert":  ("Ablation: single Expert",   "Full minus multi-agent expert split"),
    "AB_replan":       ("Ablation: no Replan",       "Full minus replan retry"),
    "AB_evaluator":    ("Ablation: no Evaluator",    "Full minus quality gate"),
}

summary = data.get("summary", {})
for exp_id, meta in EXPERIMENT_META.items():
    s = summary.get(exp_id)
    label, notes = meta
    if s is None:
        row = [label, "-", "-", "-", "-", "-", "-", notes + " (NOT RUN)"]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.fill = PARTIAL_FILL
            cell.border = BORDER
        continue
    row = [
        label,
        s.get("n", 0),
        s.get("accuracy", 0),
        s.get("faithfulness", 0),
        s.get("avg_latency", 0),
        s.get("p95_latency", 0),
        s.get("avg_tokens", 0),
        notes,
    ]
    ws.append(row)
    r = ws.max_row
    ws.cell(row=r, column=3).number_format = "0.00%"
    ws.cell(row=r, column=4).number_format = "0.000"
    ws.cell(row=r, column=5).number_format = "0.00"
    ws.cell(row=r, column=6).number_format = "0.00"
    for cell in ws[r]:
        cell.border = BORDER
        cell.alignment = LEFT if cell.column == 8 else CENTER
col_widths(ws, [26, 6, 10, 12, 16, 16, 12, 40])

# ── Sheet 2: Case-level comparison ──
ws2 = wb.create_sheet("Case Comparison")
headers2 = ["ID", "Type", "Difficulty", "Query"]
run_experiments = [eid for eid in EXPERIMENT_META if summary.get(eid) is not None]
for eid in run_experiments:
    label = EXPERIMENT_META[eid][0].split(":")[0]
    headers2 += [f"{label} score", f"{label} lat(s)"]
ws2.append(headers2)
style_header(ws2[1])

DETAILS_KEY = {
    "B0_direct_llm": "details_b0",
    "B1_dense_rag": "details_b1",
    "B2_dense_topk3": "details_b2",
    "B3_dense_topk10": "details_b3",
    "B4_no_cite_prompt": "details_b4",
    "B5_full": "details_b5",
}
details = {}
for eid in run_experiments:
    key = DETAILS_KEY.get(eid, f"details_{eid.lower()}")
    details[eid] = {r["id"]: r for r in data.get(key, [])}

case_ids = []
for d in details.values():
    for cid in d:
        if cid not in case_ids:
            case_ids.append(cid)

for cid in case_ids:
    sample = next((d[cid] for d in details.values() if cid in d), {})
    row = [cid, sample.get("type",""), sample.get("difficulty",""), sample.get("query","")[:60]]
    for eid in run_experiments:
        rec = details[eid].get(cid)
        if rec is None:
            row += ["-", "-"]
        else:
            score = rec.get("judge", {}).get("overall", 0)
            row += [score, rec.get("latency", 0)]
    ws2.append(row)
    r = ws2.max_row
    for i, eid in enumerate(run_experiments):
        col_score = 5 + i*2
        col_lat = col_score + 1
        cell = ws2.cell(row=r, column=col_score)
        cell.number_format = "0.00"
        if isinstance(cell.value, (int, float)):
            if cell.value >= 0.6:
                cell.fill = PASS_FILL
            elif cell.value >= 0.3:
                cell.fill = PARTIAL_FILL
            else:
                cell.fill = FAIL_FILL
        ws2.cell(row=r, column=col_lat).number_format = "0.0"
    for cell in ws2[r]:
        cell.border = BORDER
        cell.alignment = LEFT if cell.column == 4 else CENTER
ws2.freeze_panes = "E2"
col_widths(ws2, [10, 20, 12, 50] + [12, 10] * len(run_experiments))

# ── Sheet 3: By type breakdown ──
ws3 = wb.create_sheet("By Type")
types = ["simple_fact","recommendation","comparison","alias_resolution","multi_turn","edge_case","knowledge_boundary"]
headers3 = ["Query Type"] + [EXPERIMENT_META[eid][0].split(":")[0] + " Acc" for eid in run_experiments]
ws3.append(headers3)
style_header(ws3[1])

for t in types:
    row = [t]
    for eid in run_experiments:
        cases = [r for r in details[eid].values() if r.get("type") == t]
        if not cases:
            row.append("-")
            continue
        correct = sum(1 for r in cases if r.get("correct"))
        row.append(correct / len(cases))
    ws3.append(row)
    r = ws3.max_row
    for i in range(2, len(row)+1):
        cell = ws3.cell(row=r, column=i)
        cell.number_format = "0.00%"
        if isinstance(cell.value, float):
            if cell.value >= 0.6:
                cell.fill = PASS_FILL
            elif cell.value >= 0.3:
                cell.fill = PARTIAL_FILL
            else:
                cell.fill = FAIL_FILL
    for cell in ws3[r]:
        cell.border = BORDER
        cell.alignment = CENTER
col_widths(ws3, [22] + [14] * len(run_experiments))

# ── Sheet 4: By difficulty ──
ws4 = wb.create_sheet("By Difficulty")
diffs = ["easy","medium","hard"]
headers4 = ["Difficulty"] + [EXPERIMENT_META[eid][0].split(":")[0] + " Acc" for eid in run_experiments]
ws4.append(headers4)
style_header(ws4[1])
for d in diffs:
    row = [d]
    for eid in run_experiments:
        cases = [r for r in details[eid].values() if r.get("difficulty") == d]
        if not cases:
            row.append("-")
            continue
        correct = sum(1 for r in cases if r.get("correct"))
        row.append(correct / len(cases))
    ws4.append(row)
    r = ws4.max_row
    for i in range(2, len(row)+1):
        cell = ws4.cell(row=r, column=i)
        cell.number_format = "0.00%"
        if isinstance(cell.value, float):
            if cell.value >= 0.6:
                cell.fill = PASS_FILL
            elif cell.value >= 0.3:
                cell.fill = PARTIAL_FILL
            else:
                cell.fill = FAIL_FILL
    for cell in ws4[r]:
        cell.border = BORDER
        cell.alignment = CENTER
col_widths(ws4, [14] + [14] * len(run_experiments))

wb.save(OUT)
print(f"Saved: {OUT}")
