"""Build Hard-set sheets on top of ablation_results.xlsx.

Reads tests/hard_eval_results.json (post-rescore), computes per-run
aggregates directly from the `details_h*` rows (so results are always
consistent with the current details), then adds two sheets to the
workbook:
  Hard Summary  — H0/H1/H2 overall + per-type accuracy + latency/tokens
  Hard Cases    — case-level pass/fail + truncated answers

Any existing sheet whose title starts with "Hard" is dropped first.
"""
from __future__ import annotations

import json
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HARD_SRC = os.path.join(ROOT, "tests", "hard_eval_results.json")
XLSX = os.path.join(ROOT, "tests", "ablation_results.xlsx")

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PARTIAL_FILL = PatternFill("solid", fgColor="FFEB9C")
BORDER = Border(*[Side(style="thin", color="B4B4B4")] * 4)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(row):
    for cell in row:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> None:
    with open(HARD_SRC, encoding="utf-8") as f:
        H = json.load(f)

    DETAILS = {
        "H0": "details_h0",
        "H1": "details_h1",
        "H2": "details_h2",
        "H3": "details_h3",
        "H4": "details_h4",
        "H5": "details_h5",
        "H6": "details_h6",
        "H7": "details_h7",
    }
    # Keep only variants that actually have rows, in a stable order.
    experiments = [e for e, k in DETAILS.items() if H.get(k)]
    agg: dict[str, dict] = {}
    for exp_label in experiments:
        rows = H.get(DETAILS[exp_label], [])
        if not rows:
            continue
        n = len(rows)
        correct = sum(1 for r in rows if r["correct"])
        partial = sum(r.get("partial_score", 0.0) for r in rows) / n
        by_type: dict[str, dict] = {}
        total_lat = 0.0
        total_tok = 0
        for r in rows:
            b = by_type.setdefault(r["type"], {"n": 0, "correct": 0})
            b["n"] += 1
            b["correct"] += int(r["correct"])
            total_lat += r.get("latency", 0)
            total_tok += r.get("tokens", 0)
        for b in by_type.values():
            b["acc"] = round(b["correct"] / b["n"], 3)
        agg[exp_label] = {
            "n": n,
            "strict_accuracy": correct / n,
            "partial_score": partial,
            "by_type": by_type,
            "avg_latency": total_lat / n,
            "avg_tokens": total_tok / n,
            "is_partial_run": len(by_type) < 6,
        }

    wb = load_workbook(XLSX) if os.path.exists(XLSX) else Workbook()
    for sname in list(wb.sheetnames):
        if sname.startswith("Hard"):
            del wb[sname]

    # Ordered: original 6 first, then the 8 cold-Bangumi types added later.
    query_types = [
        "metadata_cross",
        "longtail_fact",
        "similar_recommendation",
        "bangumi_tags",
        "bangumi_score_precise",
        "kb_boundary",
        "cold_score_precise",
        "cold_longtail_fact",
        "release_date_precise",
        "cold_studio_year",
        "tag_top_score",
        "numeric_comparison",
        "seiyuu_cold_works",
        "refusal_fabricated",
    ]
    # Drop types that don't actually appear in the results (keeps the sheet tidy)
    query_types = [
        qt for qt in query_types
        if any(agg[e]["by_type"].get(qt) for e in experiments)
    ]

    ws = wb.create_sheet("Hard Summary", 0)
    ws.append(["Metric"] + experiments)
    style_header(ws[1])

    def add_row(label, values, pct=True):
        ws.append([label] + list(values))
        r = ws.max_row
        for i, v in enumerate(values, 2):
            cell = ws.cell(row=r, column=i)
            if isinstance(v, float):
                cell.number_format = "0.0%" if pct else "0.00"
                if pct:
                    if v >= 0.6:
                        cell.fill = PASS_FILL
                    elif v >= 0.3:
                        cell.fill = PARTIAL_FILL
                    else:
                        cell.fill = FAIL_FILL
        for cell in ws[r]:
            cell.border = BORDER
            cell.alignment = CENTER

    add_row(
        "Overall Strict Accuracy",
        [
            f"{agg[e]['strict_accuracy']:.1%} ({agg[e]['n']} case)"
            if agg[e]["is_partial_run"]
            else agg[e]["strict_accuracy"]
            for e in experiments
        ],
    )
    add_row(
        "Overall Partial Score",
        [
            f"{agg[e]['partial_score']:.3f} ({agg[e]['n']} case)"
            if agg[e]["is_partial_run"]
            else agg[e]["partial_score"]
            for e in experiments
        ],
        pct=False,
    )
    for qt in query_types:
        vals = []
        for e in experiments:
            b = agg[e]["by_type"].get(qt, {})
            vals.append(b.get("acc", "-"))
        add_row(qt, vals)

    add_row(
        "Avg Latency (s)",
        [agg[e]["avg_latency"] for e in experiments],
        pct=False,
    )
    add_row(
        "Avg Tokens",
        [agg[e]["avg_tokens"] for e in experiments],
        pct=False,
    )
    col_widths(ws, [28] + [12] * len(experiments))

    ws2 = wb.create_sheet("Hard Cases")
    headers = ["ID", "Type", "Query"]
    for e in experiments:
        headers += [f"{e} Correct", f"{e} Answer (truncated)"]
    ws2.append(headers)
    style_header(ws2[1])

    all_ids: list[str] = []
    for e in experiments:
        for r in H[DETAILS[e]]:
            if r["id"] not in all_ids:
                all_ids.append(r["id"])

    for cid in all_ids:
        sample = None
        for e in experiments:
            for r in H[DETAILS[e]]:
                if r["id"] == cid:
                    sample = r
                    break
            if sample:
                break
        row = [cid, sample["type"], sample["query"]]
        for e in experiments:
            key = DETAILS[e]
            rec = next((r for r in H[key] if r["id"] == cid), None)
            if rec is None:
                row += ["-", "-"]
            else:
                row += [
                    "PASS" if rec["correct"] else "FAIL",
                    (rec.get("answer") or "")[:200],
                ]
        ws2.append(row)
        r = ws2.max_row
        # column indexes of "answer" columns start at 5 and step by 2
        answer_cols = {5 + i * 2 for i in range(len(experiments))}
        query_col = 3
        for cell in ws2[r]:
            cell.border = BORDER
            cell.alignment = (
                LEFT if (cell.column == query_col or cell.column in answer_cols)
                else CENTER
            )
        for i, e in enumerate(experiments):
            col = 4 + i * 2
            cell = ws2.cell(row=r, column=col)
            if cell.value == "PASS":
                cell.fill = PASS_FILL
            elif cell.value == "FAIL":
                cell.fill = FAIL_FILL

    ws2.freeze_panes = "D2"
    ws2_widths = [10, 22, 60]
    for _ in experiments:
        ws2_widths += [10, 60]
    col_widths(ws2, ws2_widths)

    wb.save(XLSX)
    print(f"Saved: {XLSX}")
    print(f"Sheets: {wb.sheetnames}")
    for e in experiments:
        s = agg[e]
        print(
            f"  {e}: {s['strict_accuracy']:.1%}  "
            f"({s['n']} cases, {s['avg_latency']:.1f}s, {s['avg_tokens']:.0f} tok)"
        )


if __name__ == "__main__":
    main()
