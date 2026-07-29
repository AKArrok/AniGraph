"""
Build a human-labeling worksheet to measure Judge (deepseek-v4-flash) agreement.

Reads tests/ablation_results.json, picks a stratified sample of B1 cases
(balanced across judge_correct True/False), and writes tests/human_labeling.xlsx
with an empty 'human_correct' column for manual scoring.

After the human fills in 0/1 values, run tests/compute_agreement.py to get the
agreement rate + Cohen's kappa.
"""
from __future__ import annotations

import json
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).parent
RESULTS = ROOT / "ablation_results.json"
OUT = ROOT / "human_labeling.xlsx"
SAMPLE_SIZE = 20
SEED = 42


def _load_details():
    data = json.loads(RESULTS.read_text(encoding="utf-8"))
    # Prefer full B1 details; fall back to B0 if B1 not finished.
    for key in ("details_b1", "details_b1_partial", "details_b0"):
        rows = data.get(key)
        if rows:
            return key, rows
    raise SystemExit("no details in ablation_results.json")


def _stratified_sample(rows, k):
    correct = [r for r in rows if r.get("correct")]
    wrong = [r for r in rows if not r.get("correct")]
    rng = random.Random(SEED)
    rng.shuffle(correct)
    rng.shuffle(wrong)
    half = k // 2
    pick = correct[:half] + wrong[: k - half]
    if len(pick) < k:
        extras = [r for r in rows if r not in pick]
        rng.shuffle(extras)
        pick += extras[: k - len(pick)]
    rng.shuffle(pick)
    return pick


def main():
    source, rows = _load_details()
    if len(rows) < SAMPLE_SIZE:
        print(f"WARN: only {len(rows)} rows in {source}, using all")
        sample = rows
    else:
        sample = _stratified_sample(rows, SAMPLE_SIZE)

    wb = Workbook()
    ws = wb.active
    ws.title = "labeling"
    headers = [
        "id", "type", "difficulty", "query", "answer",
        "gold_keywords", "judge_overall", "judge_correct",
        "human_correct (fill 0/1)", "note",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    for r in sample:
        judge = r.get("judge") or {}
        overall = judge.get("overall", 0)
        ws.append([
            r.get("id"),
            r.get("type"),
            r.get("difficulty"),
            r.get("query", ""),
            (r.get("answer", "") or "")[:1500],
            ", ".join(r.get("gold_keywords", []) or []),
            round(float(overall), 3),
            "TRUE" if r.get("correct") else "FALSE",
            "",
            "",
        ])

    # Column widths for readability
    widths = {
        "A": 8, "B": 14, "C": 10, "D": 40, "E": 60,
        "F": 28, "G": 12, "H": 14, "I": 20, "J": 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT)
    print(f"wrote {OUT} with {len(sample)} rows sampled from {source}")
    print("Fill column I with 0 or 1, then run tests/compute_agreement.py")


if __name__ == "__main__":
    main()
