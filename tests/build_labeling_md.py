"""Emit the 20-case human labeling sheet as Markdown for in-chat labeling."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent
XLSX = ROOT / "human_labeling.xlsx"
MD = ROOT / "human_labeling.md"

if not XLSX.exists():
    raise SystemExit("run build_human_agreement.py first")

wb = load_workbook(XLSX)
ws = wb.active
header = [c.value for c in ws[1]]
rows = list(ws.iter_rows(min_row=2, values_only=True))

idx = {name: i for i, name in enumerate(header)}
lines = ["# Human labeling (fill each row with 0 or 1)\n",
         "Judge column is hidden until you decide. Only replace the `?` in `human=?`.\n"]

for i, row in enumerate(rows, 1):
    q = row[idx["query"]] or ""
    ans = (row[idx["answer"]] or "")[:1200]
    kw = row[idx["gold_keywords"]] or ""
    cid = row[idx["id"]] or ""
    typ = row[idx["type"]] or ""
    diff = row[idx["difficulty"]] or ""
    lines.append(f"---\n\n## {i}. {cid}  [{typ} / {diff}]\n")
    lines.append(f"**query**: {q}\n")
    lines.append(f"**gold keywords**: {kw}\n")
    lines.append(f"**answer**:\n\n> " + ans.replace("\n", "\n> ") + "\n")
    lines.append(f"**human = ?**  (0/1)\n")

MD.write_text("\n".join(lines), encoding="utf-8")
print(f"wrote {MD} with {len(rows)} rows")
