"""
Compute agreement between the LLM judge and the human labels filled into
tests/human_labeling.xlsx.

Reports:
  - total labeled
  - raw agreement rate
  - Cohen's kappa
  - confusion matrix (judge x human)
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent
SHEET = ROOT / "human_labeling.xlsx"


def _to_int(val):
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in {"1", "true", "t", "y", "yes", "correct"}:
        return 1
    if s in {"0", "false", "f", "n", "no", "wrong"}:
        return 0
    return None


def main():
    if not SHEET.exists():
        raise SystemExit(f"missing {SHEET}; run build_human_agreement.py first")

    wb = load_workbook(SHEET)
    ws = wb.active
    header = [c.value for c in ws[1]]
    idx_judge = header.index("judge_correct")
    idx_human = header.index("human_correct (fill 0/1)")

    pairs = []
    blanks = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        judge_raw = row[idx_judge]
        human_raw = row[idx_human]
        judge = 1 if str(judge_raw).strip().upper() == "TRUE" else 0
        human = _to_int(human_raw)
        if human is None:
            # Treat blank as 0 (user convention: unfilled = wrong).
            human = 0
            blanks += 1
        pairs.append((judge, human))
    if blanks:
        print(f"(treated {blanks} blank rows as human=0)")

    if not pairs:
        raise SystemExit("no human labels filled in yet")

    n = len(pairs)
    agree = sum(1 for j, h in pairs if j == h)
    p_o = agree / n

    j1 = sum(j for j, _ in pairs) / n
    h1 = sum(h for _, h in pairs) / n
    p_e = j1 * h1 + (1 - j1) * (1 - h1)
    kappa = (p_o - p_e) / (1 - p_e) if p_e != 1 else float("nan")

    tp = sum(1 for j, h in pairs if j == 1 and h == 1)
    tn = sum(1 for j, h in pairs if j == 0 and h == 0)
    fp = sum(1 for j, h in pairs if j == 1 and h == 0)
    fn = sum(1 for j, h in pairs if j == 0 and h == 1)

    print(f"labeled: {n}")
    print(f"raw agreement: {p_o:.3f}  ({agree}/{n})")
    print(f"Cohen's kappa: {kappa:.3f}")
    print(f"judge=T,human=T (TP): {tp}")
    print(f"judge=T,human=F (FP over-credit): {fp}")
    print(f"judge=F,human=T (FN under-credit): {fn}")
    print(f"judge=F,human=F (TN): {tn}")
    if kappa >= 0.8:
        print("kappa >= 0.8: judge is trustworthy.")
    elif kappa >= 0.6:
        print("kappa in [0.6,0.8): substantial agreement; usable with caveats.")
    else:
        print("kappa < 0.6: judge unreliable, revisit prompt or model.")


if __name__ == "__main__":
    main()
