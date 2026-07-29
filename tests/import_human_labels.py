"""Read a list of 20 human labels (space/comma/newline separated 0/1) and
write them into column I of tests/human_labeling.xlsx.

Usage:
    python tests/import_human_labels.py "1 0 1 1 0 ..."
or pipe:
    echo 1 0 1 1 0 ... | python tests/import_human_labels.py
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent
XLSX = ROOT / "human_labeling.xlsx"


def _parse(text: str) -> list[int]:
    text = text.replace(",", " ").replace("\n", " ")
    out: list[int] = []
    for tok in text.split():
        tok = tok.strip()
        if tok == "":
            continue
        if tok not in {"0", "1"}:
            raise SystemExit(f"invalid token: {tok!r} (only 0/1 allowed)")
        out.append(int(tok))
    return out


def main() -> None:
    if len(sys.argv) > 1:
        raw = " ".join(sys.argv[1:])
    else:
        raw = sys.stdin.read()
    labels = _parse(raw)

    wb = load_workbook(XLSX)
    ws = wb.active
    header = [c.value for c in ws[1]]
    col_idx = header.index("human_correct (fill 0/1)") + 1
    n_rows = ws.max_row - 1
    if len(labels) != n_rows:
        raise SystemExit(f"got {len(labels)} labels but sheet has {n_rows} rows")

    for i, lbl in enumerate(labels, start=2):
        ws.cell(row=i, column=col_idx, value=lbl)
    wb.save(XLSX)
    print(f"wrote {len(labels)} labels into {XLSX}")


if __name__ == "__main__":
    main()
